#Import all the required packages
import requests
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, render_template
import time
# Flask constructor takes the name of current module (__name__) as argument
app = Flask(__name__)
@app.route('/', methods=("POST", "GET"))
#Define a function with name execute
def execute():
    print("start")
    #Read the data from the sheet2 of excel file where the input resides, into the pandas dataframe
    excel_data_df = pd.read_excel('webAPI.xlsx', sheet_name='Sheet2')
    #Add one more column to the dataframe
    excel_data_df.loc[excel_data_df.Unit == 'F', 'Unit1'] = 'imperial'
    excel_data_df.loc[excel_data_df.Unit == 'C', 'Unit1'] = 'Metric'
    excel_data_df.loc[excel_data_df.Unit == 'K', 'Unit1'] = 'Kelvin'
    #Filter the data based on the Update(0/1) column
    update_Df = excel_data_df[excel_data_df['Update'] == 1]
    non_update_Df = excel_data_df[excel_data_df['Update'] == 0]
    #Set the base_url and api_key in the variables for future use
    BASE_URL = "https://api.openweathermap.org/data/2.5/weather?"
    API_KEY = "5979f663970b91b0d438ff51c76917a2"
    dd = {}
    # Define a function with dataframe as a parameter
    def update(df):
        t=[]
        h=[]
        for ind in df.index:
            #Pass the cityname and units columns from the dataframe, to the url along with the api_key
            URL = BASE_URL + "q=" + df['CityName'][ind] + "&units="+df['Unit1'][ind]+"&appid=" + API_KEY
            # get method of requests module to return response object
            response = requests.get(URL)
            #json method of response object .convert json format data into python format data
            main = response.json()['main']
            City = df['CityName'][ind]
            # store the value corresponding to the "temp" and 'humidity' key of main
            humidity = main['humidity']
            temperature = main['temp']
            t.append(temperature)
            h.append(humidity)
            
            #Add the necessary columns into the dictionary which needs to be updated in web page
            dd[City] = {'City': City, 'Temp': temperature, 'humidity': humidity,'Unit':df['Unit'][ind],'Update':df['Update'][ind]}
        df.insert(1, "temperature", t, True)
        df.insert(2, "humidity", h, True)
        return df
    #Call the update function with non_update dataframe
    a=update(non_update_Df)
    #Concatenate the non_update dataframe with the update_dataframe into the result dataframe
    result = pd.concat([a, update(update_Df)])
    #Remove the unnecessary columns in the dataframe
    result=result.drop(columns="Unit1")
    #Print the resulted dataframe
    print(result)
    #Save the result dataframe into the sheet1 of same excel file(which we read in the beginning) with the help of load_workbook
    book = load_workbook('webAPI.xlsx')
    writer = pd.ExcelWriter('webAPI.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    result.to_excel(writer, "Sheet1",index = False)
    writer.save()
    #Set the timer to update the records for every 2 seconds
    time.sleep(2)
    #Use the render_template() method to render the TempApp.html
    return render_template('TempApp.html',  result=dd)
    #Make the function to run itertively
    while True:
        execute()
# main driver function
if __name__ == '__main__':
     app.run(debug = True)
